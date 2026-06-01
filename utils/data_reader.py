import csv
import json
from pathlib import Path
from typing import List, Dict, Any


class DataReader:
    @staticmethod
    def read_json(file_path: str) -> Dict[str, Any]:
        path = Path(file_path)
        with path.open('r', encoding='utf-8') as json_file:
            return json.load(json_file)

    @staticmethod
    def read_csv(file_path: str) -> List[Dict[str, str]]:
        path = Path(file_path)
        with path.open('r', encoding='utf-8') as csv_file:
            return [row for row in csv.DictReader(csv_file)]

    @staticmethod
    def read_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name or workbook.sheetnames[0]]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        return [
            {headers[index]: cell.value for index, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ]
